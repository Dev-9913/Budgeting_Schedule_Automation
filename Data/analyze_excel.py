import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
import json

def analyze_excel_structure(filename):
    """Comprehensive analysis of the budgeting Excel file"""

    wb = openpyxl.load_workbook(filename, data_only=False)
    wb_data = openpyxl.load_workbook(filename, data_only=True)

    analysis = {
        'total_sheets': len(wb.sheetnames),
        'sheet_names': wb.sheetnames,
        'sheets_detail': {}
    }

    print("="*80)
    print(f"COMPREHENSIVE EXCEL ANALYSIS: {filename}")
    print("="*80)
    print(f"\nTotal Sheets: {len(wb.sheetnames)}")
    print(f"Sheet Names: {', '.join(wb.sheetnames)}\n")

    for sheet_name in wb.sheetnames:
        print("\n" + "="*80)
        print(f"SHEET: {sheet_name}")
        print("="*80)

        ws = wb[sheet_name]
        ws_data = wb_data[sheet_name]

        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column

        print(f"Dimensions: {max_row} rows x {max_col} columns")

        # Analyze formulas vs values
        formulas = []
        input_cells = []
        merged_cells = list(ws.merged_cells.ranges)

        for row in range(1, min(max_row + 1, 100)):  # Limit to first 100 rows
            for col in range(1, max_col + 1):
                cell = ws.cell(row, col)
                cell_ref = f"{get_column_letter(col)}{row}"

                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formulas.append({
                            'cell': cell_ref,
                            'formula': cell.value,
                            'result': ws_data.cell(row, col).value
                        })
                    elif not isinstance(cell.value, str) or not cell.value.startswith('='):
                        # This might be an input cell
                        input_cells.append({
                            'cell': cell_ref,
                            'value': cell.value,
                            'type': type(cell.value).__name__
                        })

        print(f"\nFormula cells found: {len(formulas)}")
        print(f"Potential input cells: {len(input_cells)}")
        print(f"Merged cell ranges: {len(merged_cells)}")

        # Show first few rows of data
        print("\nFirst 10 rows preview:")
        print("-" * 80)
        for row in range(1, min(11, max_row + 1)):
            row_data = []
            for col in range(1, min(max_col + 1, 10)):  # First 10 columns
                cell_value = ws.cell(row, col).value
                if cell_value is None:
                    row_data.append("")
                elif isinstance(cell_value, str) and cell_value.startswith('='):
                    row_data.append(f"[FORMULA]")
                else:
                    row_data.append(str(cell_value)[:20])  # Truncate long values
            print(f"Row {row:2d}: {' | '.join(row_data)}")

        # Show sample formulas
        if formulas:
            print("\nSample formulas (first 10):")
            print("-" * 80)
            for f in formulas[:10]:
                print(f"{f['cell']}: {f['formula'][:100]}")

        # Identify reference patterns
        external_refs = [f for f in formulas if '!' in f['formula']]
        if external_refs:
            print(f"\nCross-sheet references found: {len(external_refs)}")
            print("Sample cross-sheet references:")
            for ref in external_refs[:5]:
                print(f"  {ref['cell']}: {ref['formula'][:80]}")

        analysis['sheets_detail'][sheet_name] = {
            'dimensions': f"{max_row}x{max_col}",
            'formula_count': len(formulas),
            'input_count': len(input_cells),
            'has_cross_refs': len(external_refs) > 0
        }

    print("\n" + "="*80)
    print("SUMMARY")
    print("="*80)
    print(json.dumps(analysis, indent=2))

    return analysis

if __name__ == "__main__":
    analyze_excel_structure("Management Accounting 13 Schedules Excel.xlsx")

